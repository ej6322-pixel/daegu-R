"""
백화점 경쟁사 행사 AI 분석 시스템 — Streamlit 버전
Streamlit Cloud 배포용 (이미지 업로드 기반)
"""
import os, base64, io
from datetime import datetime
import streamlit as st
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── API 키 설정 ──
def get_api_key():
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
        if key:
            return key
    except Exception:
        pass
    return os.environ.get("ANTHROPIC_API_KEY", "")

SYSTEM = "당신은 백화점 MD 경쟁분석 전문가입니다."

# Tool Use 스키마 — extract_events용
EXTRACT_TOOL = [{
    "name": "save_events",
    "description": "이미지에서 추출한 행사 목록을 저장합니다.",
    "input_schema": {
        "type": "object",
        "properties": {
            "events": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "category": {"type": "string"},
                        "name":     {"type": "string"},
                        "detail":   {"type": "string"},
                        "period":   {"type": "string"},
                        "type":     {"type": "string"},
                    },
                    "required": ["category", "name", "detail", "period", "type"],
                },
            }
        },
        "required": ["events"],
    },
}]

# Tool Use 스키마 — compare용
COMPARE_TOOL = [{
    "name": "save_analysis",
    "description": "두 백화점 행사 비교 분석 결과를 저장합니다.",
    "input_schema": {
        "type": "object",
        "properties": {
            "categories": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "category": {"type": "string"},
                        "lotte":    {"type": "string"},
                        "hyundai":  {"type": "string"},
                        "winner":   {"type": "string", "enum": ["롯데", "더현대", "비슷"]},
                        "point":    {"type": "string"},
                    },
                    "required": ["category", "lotte", "hyundai", "winner", "point"],
                },
            },
            "saeunn": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "type":    {"type": "string"},
                        "lotte":   {"type": "string"},
                        "hyundai": {"type": "string"},
                        "winner":  {"type": "string", "enum": ["롯데", "더현대", "비슷"]},
                        "point":   {"type": "string"},
                    },
                    "required": ["type", "lotte", "hyundai", "winner", "point"],
                },
            },
            "lotte_strength":   {"type": "array", "items": {"type": "string"}},
            "hyundai_strength": {"type": "array", "items": {"type": "string"}},
            "insight":          {"type": "array", "items": {"type": "string"}},
        },
        "required": ["categories", "saeunn", "lotte_strength", "hyundai_strength", "insight"],
    },
}]


def extract_events(cl, uploaded_files, store_name, model="claude-sonnet-4-6"):
    if not uploaded_files:
        return []
    content = []
    for f in uploaded_files:
        data = base64.b64encode(f.read()).decode()
        mime = f.type or "image/jpeg"
        content.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": data}})
        f.seek(0)
    content.append({"type": "text", "text": (
        f"이 이미지들은 {store_name} 카카오채널 스크린샷입니다. "
        "모든 행사·팝업·사은혜택·이벤트를 빠짐없이 추출하고 save_events 툴로 저장하세요. "
        "category 분류: 패션/스포츠·레저/뷰티/식품F&B/리빙가구/팝업스토어/사은혜택/문화이벤트"
    )})
    resp = cl.messages.create(
        model=model, max_tokens=4096,
        system=SYSTEM,
        tools=EXTRACT_TOOL,
        tool_choice={"type": "tool", "name": "save_events"},
        messages=[{"role": "user", "content": content}],
    )
    for block in resp.content:
        if block.type == "tool_use":
            return block.input.get("events", [])
    return []


def compare(cl, lotte, hyundai, model="claude-sonnet-4-6"):
    lt = "\n".join([f"[{e['category']}] {e['name']}: {e['detail']}" for e in lotte])
    ht = "\n".join([f"[{e['category']}] {e['name']}: {e['detail']}" for e in hyundai])
    prompt = (
        "롯데백화점 대구점과 더현대 대구의 행사를 비교 분석하고 save_analysis 툴로 저장하세요.\n\n"
        f"[롯데 행사]\n{lt or '(정보 없음)'}\n\n"
        f"[더현대 행사]\n{ht or '(정보 없음)'}\n\n"
        "categories: 패션/스포츠·레저/뷰티/식품F&B/리빙가구/팝업스토어/문화이벤트 별로 작성\n"
        "saeunn: 사은품·경품/추가할인·쿠폰/적립혜택/VIP혜택/제휴카드혜택/기타 별로 작성\n"
        "insight: 롯데 영업기획팀 입장의 실전 전략 제언 4가지"
    )
    resp = cl.messages.create(
        model=model, max_tokens=4096,
        tools=COMPARE_TOOL,
        tool_choice={"type": "tool", "name": "save_analysis"},
        messages=[{"role": "user", "content": prompt}],
    )
    for block in resp.content:
        if block.type == "tool_use":
            return block.input
    raise ValueError("AI가 분석 결과를 반환하지 않았습니다. 다시 시도해주세요.")


def build_excel(data):
    wb = Workbook()

    def th(c, bg="1A1A2E", fc="FFFFFF", sz=10):
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(bold=True, color=fc, name="맑은 고딕", size=sz)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        s = Side(border_style="thin", color="CCCCCC")
        c.border = Border(left=s, right=s, top=s, bottom=s)

    def td(c, bg="FFFFFF", fc="111111", bold=False, center=False):
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(bold=bold, color=fc, name="맑은 고딕", size=9)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
        s = Side(border_style="thin", color="E0E0E0")
        c.border = Border(left=s, right=s, top=s, bottom=s)

    # 시트1: 상품군별 비교
    ws1 = wb.active; ws1.title = "상품군별 비교"; ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:E1")
    c = ws1["A1"]; c.value = f"롯데 vs 더현대 상품군별 행사 비교 ({data.get('analyzed_at', '')})"; th(c, sz=12)
    ws1.row_dimensions[1].height = 26
    for i, (h, w) in enumerate([("상품군", 14), ("롯데백화점", 36), ("더현대 대구", 36), ("우세", 10), ("비교포인트", 32)], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
        th(ws1.cell(row=2, column=i, value=h))
    for r, cat in enumerate(data.get("categories", []), 3):
        ws1.row_dimensions[r].height = 42
        bg = "FAFAFA" if r % 2 == 0 else "FFFFFF"
        w = cat.get("winner", "비슷"); wfc = "C00020" if w == "롯데" else ("003087" if w == "더현대" else "555555")
        for ci, v in enumerate([cat.get("category"), cat.get("lotte", "—"), cat.get("hyundai", "—"), w, cat.get("point", "")], 1):
            td(ws1.cell(row=r, column=ci, value=v), bg=bg, bold=(ci == 1 or ci == 4), fc=wfc if ci == 4 else "111111", center=(ci == 1 or ci == 4))

    # 시트2: 사은행사 비교
    ws_sa = wb.create_sheet("사은행사 비교"); ws_sa.sheet_view.showGridLines = False
    ws_sa.merge_cells("A1:E1"); c = ws_sa["A1"]; c.value = f"롯데 vs 더현대 사은행사 비교 ({data.get('analyzed_at', '')})"; th(c, bg="856404", sz=12); ws_sa.row_dimensions[1].height = 26
    for i, (h, w) in enumerate([("사은행사 유형", 18), ("롯데백화점", 36), ("더현대 대구", 36), ("우세", 10), ("비교포인트", 32)], 1):
        ws_sa.column_dimensions[get_column_letter(i)].width = w
        th(ws_sa.cell(row=2, column=i, value=h), bg="6b4f00")
    for r, sa in enumerate(data.get("saeunn", []), 3):
        ws_sa.row_dimensions[r].height = 42
        bg = "FFFDE7" if r % 2 == 0 else "FFFFFF"
        w = sa.get("winner", "비슷"); wfc = "C00020" if w == "롯데" else ("003087" if w == "더현대" else "555555")
        for ci, v in enumerate([sa.get("type"), sa.get("lotte", "—"), sa.get("hyundai", "—"), w, sa.get("point", "")], 1):
            td(ws_sa.cell(row=r, column=ci, value=v), bg=bg, bold=(ci == 1 or ci == 4), fc=wfc if ci == 4 else "111111", center=(ci == 1 or ci == 4))

    # 시트3: 롯데 행사 상세
    ws2 = wb.create_sheet("롯데 행사 상세"); ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:E1"); c = ws2["A1"]; c.value = "롯데백화점 대구점 행사 전체"; th(c, bg="E8002D", sz=12); ws2.row_dimensions[1].height = 24
    for i, (h, w) in enumerate([("카테고리", 12), ("행사명", 28), ("내용", 42), ("기간", 14), ("유형", 10)], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w; th(ws2.cell(row=2, column=i, value=h), bg="C00020")
    for r, ev in enumerate(data.get("lotte_events", []), 3):
        bg = "FFF5F5" if r % 2 == 0 else "FFFFFF"
        for ci, k in enumerate(["category", "name", "detail", "period", "type"], 1):
            td(ws2.cell(row=r, column=ci, value=ev.get(k, "")), bg=bg)
        ws2.row_dimensions[r].height = 36

    # 시트4: 더현대 행사 상세
    ws3 = wb.create_sheet("더현대 행사 상세"); ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:E1"); c = ws3["A1"]; c.value = "더현대 대구 행사 전체"; th(c, bg="003087", sz=12); ws3.row_dimensions[1].height = 24
    for i, (h, w) in enumerate([("카테고리", 12), ("행사명", 28), ("내용", 42), ("기간", 14), ("유형", 10)], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w; th(ws3.cell(row=2, column=i, value=h), bg="003087")
    for r, ev in enumerate(data.get("hyundai_events", []), 3):
        bg = "F0F4FF" if r % 2 == 0 else "FFFFFF"
        for ci, k in enumerate(["category", "name", "detail", "period", "type"], 1):
            td(ws3.cell(row=r, column=ci, value=ev.get(k, "")), bg=bg)
        ws3.row_dimensions[r].height = 36

    # 시트5: AI 분석
    ws4 = wb.create_sheet("AI 분석"); ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 18; ws4.column_dimensions["B"].width = 60
    ws4.merge_cells("A1:B1"); c = ws4["A1"]; c.value = "AI 경쟁 분석 — MD 전략 제언"; th(c, sz=12); ws4.row_dimensions[1].height = 24
    row = 2
    for title, items, bg, fc in [
        ("롯데 강점", data.get("lotte_strength", []), "FCE4E4", "C00020"),
        ("더현대 강점", data.get("hyundai_strength", []), "E4EBF9", "003087"),
        ("MD 전략 제언", data.get("insight", []), "FFFDE7", "856404"),
    ]:
        th(ws4.cell(row=row, column=1, value=title), bg=fc)
        ws4.cell(row=row, column=2).fill = PatternFill("solid", start_color=fc)
        ws4.row_dimensions[row].height = 20; row += 1
        for item in items:
            td(ws4.cell(row=row, column=2, value=f"• {item}"), bg=bg)
            ws4.cell(row=row, column=1).fill = PatternFill("solid", start_color=bg)
            ws4.row_dimensions[row].height = 30; row += 1
        row += 1

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="백화점 행사 AI 분석", page_icon="🏬", layout="wide")

# 사이드바
preset_key = get_api_key()
api_key = st.sidebar.text_input(
    "🔑 Anthropic API Key",
    value=preset_key,
    type="password",
    placeholder="sk-ant-...",
    help="console.anthropic.com 에서 발급",
)
if api_key:
    st.sidebar.success("✅ API 키 설정됨")
else:
    st.sidebar.warning("⚠️ API 키를 입력해주세요")

st.sidebar.divider()
model = st.sidebar.selectbox(
    "AI 모델",
    ["claude-sonnet-4-6", "claude-opus-4-7"],
)
st.sidebar.divider()
with st.sidebar.expander("💡 사용 가이드"):
    st.markdown("""
1. API 키 입력
2. 롯데 / 더현대 스크린샷 업로드
3. AI 분석 시작 클릭
4. 결과 확인 후 Excel 다운로드
    """)
st.sidebar.caption("🏬 백화점 행사 AI 분석 v3.0")

# 메인
st.markdown("## 🏬 백화점 행사 AI 분석")
st.caption("롯데백화점 대구점 vs 더현대 대구 — 카카오채널 스크린샷 업로드")

col_l, col_h = st.columns(2)
with col_l:
    st.markdown("### 🔴 롯데백화점 대구점")
    lotte_files = st.file_uploader(
        "카카오채널 스크린샷 (최대 10장)",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        key="lotte",
    )
    if lotte_files:
        cols = st.columns(min(len(lotte_files), 5))
        for i, f in enumerate(lotte_files[:10]):
            cols[i % 5].image(f, width=80)

with col_h:
    st.markdown("### 🔵 더현대 대구")
    hyundai_files = st.file_uploader(
        "카카오채널 스크린샷 (최대 10장)",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        key="hyundai",
    )
    if hyundai_files:
        cols = st.columns(min(len(hyundai_files), 5))
        for i, f in enumerate(hyundai_files[:10]):
            cols[i % 5].image(f, width=80)

st.divider()

if st.button("🤖 AI 분석 시작", type="primary", use_container_width=True):
    if not lotte_files and not hyundai_files:
        st.error("이미지를 최소 한 장 이상 업로드해주세요.")
    elif not api_key:
        st.error("왼쪽 사이드바에 Anthropic API 키를 입력해주세요.")
    else:
        cl = anthropic.Anthropic(api_key=api_key)
        with st.spinner("이미지 분석 중... (20~40초 소요)"):
            try:
                lotte_ev = extract_events(cl, lotte_files[:10], "롯데백화점 대구점", model)
                hyundai_ev = extract_events(cl, hyundai_files[:10], "더현대 대구", model)
                result = compare(cl, lotte_ev, hyundai_ev, model)
                result["lotte_events"] = lotte_ev
                result["hyundai_events"] = hyundai_ev
                result["analyzed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                st.session_state["result"] = result
                st.success("✅ 분석 완료!")
            except Exception as e:
                st.error(f"분석 실패: {e}")

if "result" in st.session_state:
    result = st.session_state["result"]

    tab1, tab2 = st.tabs(["🛍️ 상품군별 행사 비교", "🎁 사은행사 비교"])

    with tab1:
        categories = result.get("categories", [])
        if categories:
            rows = []
            for c in categories:
                w = c.get("winner", "비슷")
                badge = "🔴 롯데 우세" if w == "롯데" else ("🔵 더현대 우세" if w == "더현대" else "⚪ 비슷")
                rows.append({
                    "상품군": c.get("category", ""),
                    "롯데백화점": c.get("lotte", "—"),
                    "더현대 대구": c.get("hyundai", "—"),
                    "우세": badge,
                    "비교 포인트": c.get("point", ""),
                })
            st.dataframe(rows, use_container_width=True, hide_index=True)
        else:
            st.info("상품군별 행사 정보가 없습니다.")

    with tab2:
        saeunn = result.get("saeunn", [])
        if saeunn:
            rows = []
            for s in saeunn:
                w = s.get("winner", "비슷")
                badge = "🔴 롯데 우세" if w == "롯데" else ("🔵 더현대 우세" if w == "더현대" else "⚪ 비슷")
                rows.append({
                    "사은행사 유형": s.get("type", ""),
                    "롯데백화점": s.get("lotte", "—"),
                    "더현대 대구": s.get("hyundai", "—"),
                    "우세": badge,
                    "비교 포인트": s.get("point", ""),
                })
            st.dataframe(rows, use_container_width=True, hide_index=True)
        else:
            st.info("사은행사 정보가 없습니다.")

    st.subheader("🧠 AI 경쟁 분석")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("**🔴 롯데 강점**")
        for x in result.get("lotte_strength", []):
            st.markdown(f"— {x}")
    with c2:
        st.markdown("**🔵 더현대 강점**")
        for x in result.get("hyundai_strength", []):
            st.markdown(f"— {x}")
    with c3:
        st.markdown("**⚡ MD 전략 제언**")
        for x in result.get("insight", []):
            st.markdown(f"— {x}")

    st.divider()
    excel_buf = build_excel(result)
    fname = f"백화점행사비교_{datetime.now().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="📥 Excel 다운로드",
        data=excel_buf,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
