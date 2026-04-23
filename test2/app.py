"""
백화점 경쟁사 행사 AI 분석 시스템 — app.py
실행: pip install flask anthropic openpyxl pillow
      export ANTHROPIC_API_KEY="sk-ant-..."
      python app.py  →  http://localhost:5000
"""
import os, json, base64, io, re
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template_string
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY",""))

HTML = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>백화점 행사 AI 분석</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{--lotte:#e8002d;--hyundai:#003087;--bg:#f4f3f0;--card:#fff;--text:#111;--muted:#777;--border:#e0e0e0;--radius:10px}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Noto Sans KR',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
header{background:#111;color:#fff;padding:1.25rem 2rem;display:flex;align-items:center;gap:1rem}
header h1{font-size:1rem;font-weight:500;letter-spacing:-.02em}
header .badge{font-family:'DM Mono',monospace;font-size:.7rem;background:#333;padding:.25rem .6rem;border-radius:4px;color:#aaa}
.container{max-width:1160px;margin:0 auto;padding:2rem 1.5rem}
.upload-row{display:grid;grid-template-columns:1fr 1fr;gap:1.25rem;margin-bottom:1.25rem}
.ucard{background:var(--card);border:1.5px solid var(--border);border-radius:var(--radius);overflow:hidden}
.ucard-head{padding:.9rem 1.1rem;display:flex;align-items:center;gap:.6rem;border-bottom:1px solid var(--border)}
.dot{width:10px;height:10px;border-radius:50%}.dot-lotte{background:var(--lotte)}.dot-hyundai{background:var(--hyundai)}
.ucard-head span{font-weight:700;font-size:.85rem}.ucard-head small{margin-left:auto;color:var(--muted);font-size:.75rem}
.drop-zone{margin:1rem;border:1.5px dashed var(--border);border-radius:8px;padding:1.5rem 1rem;text-align:center;cursor:pointer;transition:border-color .2s,background .2s;min-height:110px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:.4rem}
.drop-zone:hover,.drop-zone.drag{border-color:#888;background:#fafafa}
.drop-zone input{display:none}.drop-zone .icon{font-size:1.6rem}.drop-zone p{font-size:.8rem;color:var(--muted)}
.preview-list{display:flex;flex-wrap:wrap;gap:.5rem;padding:0 1rem 1rem}
.preview-list img{width:60px;height:60px;object-fit:cover;border-radius:6px;border:1px solid var(--border)}
.preview-list .del-wrap{position:relative;cursor:pointer}
.preview-list .del-wrap::after{content:'✕';position:absolute;top:-4px;right:-4px;background:#111;color:#fff;font-size:.6rem;width:14px;height:14px;border-radius:50%;display:flex;align-items:center;justify-content:center}
.btn-row{display:flex;gap:.75rem;margin-bottom:1.5rem}
.btn{padding:.75rem 1.75rem;border:none;border-radius:8px;font-family:'Noto Sans KR',sans-serif;font-size:.9rem;font-weight:500;cursor:pointer;transition:opacity .2s,transform .1s}
.btn:active{transform:scale(.98)}.btn-primary{background:#111;color:#fff}.btn-primary:hover{opacity:.85}
.btn-excel{background:#1d6f42;color:#fff}.btn-excel:hover{opacity:.85}.btn-excel:disabled{background:#aaa;cursor:not-allowed}
#loading{display:none;align-items:center;gap:.75rem;padding:1rem;color:var(--muted);font-size:.85rem}
.spinner{width:20px;height:20px;border:2px solid #ddd;border-top-color:#111;border-radius:50%;animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
#result{display:none}
.section-title{font-size:.7rem;font-family:'DM Mono',monospace;color:var(--muted);letter-spacing:.08em;text-transform:uppercase;margin-bottom:.75rem}
.compare-wrap{overflow-x:auto;margin-bottom:1.5rem}
table{width:100%;border-collapse:collapse;font-size:.82rem;background:var(--card);border-radius:var(--radius);overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.06)}
thead tr{background:#111;color:#fff}
thead th{padding:.7rem .9rem;text-align:left;font-weight:500;white-space:nowrap}
tbody tr{border-bottom:1px solid var(--border)}tbody tr:last-child{border-bottom:none}
tbody td{padding:.65rem .9rem;vertical-align:top;line-height:1.5}
tbody tr:nth-child(even){background:#fafafa}
.analysis-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:1rem;margin-bottom:1.5rem}
.acard{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);padding:1.1rem}
.acard h3{font-size:.78rem;font-weight:700;margin-bottom:.6rem}
.acard ul{list-style:none;font-size:.8rem;color:#333;display:flex;flex-direction:column;gap:.35rem}
.acard ul li::before{content:'— ';color:#aaa}
.strength-l{border-top:3px solid var(--lotte)}.strength-h{border-top:3px solid var(--hyundai)}.insight{border-top:3px solid #f0a500}
.badge-l{color:var(--lotte);font-weight:700}.badge-h{color:var(--hyundai);font-weight:700}
@media(max-width:700px){.upload-row{grid-template-columns:1fr}.analysis-grid{grid-template-columns:1fr}}
</style>
</head>
<body>
<header>
  <h1>백화점 행사 AI 분석</h1>
  <span class="badge">롯데 vs 더현대</span>
</header>
<div class="container">
  <div class="upload-row">
    <div class="ucard">
      <div class="ucard-head"><div class="dot dot-lotte"></div><span>롯데백화점 대구점</span><small id="cnt-lotte">0장</small></div>
      <div class="drop-zone" id="drop-lotte" onclick="document.getElementById('inp-lotte').click()">
        <input type="file" id="inp-lotte" accept="image/*" multiple onchange="addFiles('lotte',this.files)">
        <div class="icon">📲</div><p>카카오채널 스크린샷 업로드<br>여러 장 동시 선택 가능</p>
      </div>
      <div class="preview-list" id="prev-lotte"></div>
    </div>
    <div class="ucard">
      <div class="ucard-head"><div class="dot dot-hyundai"></div><span>더현대 대구</span><small id="cnt-hyundai">0장</small></div>
      <div class="drop-zone" id="drop-hyundai" onclick="document.getElementById('inp-hyundai').click()">
        <input type="file" id="inp-hyundai" accept="image/*" multiple onchange="addFiles('hyundai',this.files)">
        <div class="icon">📲</div><p>카카오채널 스크린샷 업로드<br>여러 장 동시 선택 가능</p>
      </div>
      <div class="preview-list" id="prev-hyundai"></div>
    </div>
  </div>
  <div class="btn-row">
    <button class="btn btn-primary" onclick="analyze()">AI 분석 시작</button>
    <button class="btn btn-excel" id="btn-excel" disabled onclick="downloadExcel()">Excel 다운로드</button>
  </div>
  <div id="loading"><div class="spinner"></div>이미지 분석 중... (20~40초 소요)</div>
  <div id="result">
    <p class="section-title">상품군별 비교표</p>
    <div class="compare-wrap"><table id="tbl"></table></div>
    <p class="section-title">AI 경쟁 분석</p>
    <div class="analysis-grid" id="ag"></div>
  </div>
</div>
<script>
const files={lotte:[],hyundai:[]};
let lastResult=null;
function addFiles(s,fl){for(const f of fl){if(files[s].length>=10)break;files[s].push(f);}renderPrev(s);}
function renderPrev(s){
  const el=document.getElementById('prev-'+s);
  document.getElementById('cnt-'+s).textContent=files[s].length+'장';
  el.innerHTML='';
  files[s].forEach((f,i)=>{
    const url=URL.createObjectURL(f);
    const w=document.createElement('div');w.className='del-wrap';
    w.onclick=()=>{files[s].splice(i,1);renderPrev(s);};
    const img=document.createElement('img');img.src=url;
    w.appendChild(img);el.appendChild(w);
  });
}
['lotte','hyundai'].forEach(s=>{
  const z=document.getElementById('drop-'+s);
  z.addEventListener('dragover',e=>{e.preventDefault();z.classList.add('drag');});
  z.addEventListener('dragleave',()=>z.classList.remove('drag'));
  z.addEventListener('drop',e=>{e.preventDefault();z.classList.remove('drag');addFiles(s,e.dataTransfer.files);});
});
async function analyze(){
  if(!files.lotte.length&&!files.hyundai.length){alert('이미지를 업로드해주세요.');return;}
  document.getElementById('loading').style.display='flex';
  document.getElementById('result').style.display='none';
  document.getElementById('btn-excel').disabled=true;
  const fd=new FormData();
  for(const f of files.lotte)fd.append('lotte',f);
  for(const f of files.hyundai)fd.append('hyundai',f);
  try{
    const res=await fetch('/analyze',{method:'POST',body:fd});
    const data=await res.json();
    if(data.error){alert(data.error);return;}
    lastResult=data;renderResult(data);
    document.getElementById('btn-excel').disabled=false;
  }catch(e){alert('분석 실패: '+e.message);}
  finally{document.getElementById('loading').style.display='none';}
}
function renderResult(data){
  const tbl=document.getElementById('tbl');
  let html='<thead><tr><th>상품군</th><th style="color:#ffb3b3">롯데백화점</th><th style="color:#99b3ff">더현대 대구</th><th>우세</th><th>비교 포인트</th></tr></thead><tbody>';
  for(const c of data.categories){
    const w=c.winner==='롯데'?'<span class="badge-l">롯데 우세</span>':c.winner==='더현대'?'<span class="badge-h">더현대 우세</span>':'비슷';
    html+=`<tr><td><strong>${c.category}</strong></td><td>${c.lotte||'<span style="color:#bbb">—</span>'}</td><td>${c.hyundai||'<span style="color:#bbb">—</span>'}</td><td>${w}</td><td><small style="color:#777">${c.point||''}</small></td></tr>`;
  }
  tbl.innerHTML=html+'</tbody>';
  document.getElementById('ag').innerHTML=`
    <div class="acard strength-l"><h3><span style="color:var(--lotte)">●</span> 롯데 강점</h3><ul>${data.lotte_strength.map(x=>'<li>'+x+'</li>').join('')}</ul></div>
    <div class="acard strength-h"><h3><span style="color:var(--hyundai)">●</span> 더현대 강점</h3><ul>${data.hyundai_strength.map(x=>'<li>'+x+'</li>').join('')}</ul></div>
    <div class="acard insight"><h3>⚡ MD 전략 제언</h3><ul>${data.insight.map(x=>'<li>'+x+'</li>').join('')}</ul></div>`;
  document.getElementById('result').style.display='block';
}
async function downloadExcel(){
  if(!lastResult)return;
  const res=await fetch('/excel',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(lastResult)});
  const blob=await res.blob();
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');a.href=url;
  a.download=`백화점행사비교_${new Date().toISOString().slice(0,10)}.xlsx`;a.click();
}
</script>
</body></html>"""

SYSTEM = "당신은 백화점 MD 경쟁분석 전문가입니다. 이미지에서 행사 정보를 추출해 JSON으로만 응답하세요. 마크다운 없이 순수 JSON만 출력하세요."

def extract_events(img_files, store_name):
    if not img_files: return []
    content = []
    for f in img_files:
        data = base64.b64encode(f.read()).decode()
        content.append({"type":"image","source":{"type":"base64","media_type":f.content_type or "image/jpeg","data":data}})
    content.append({"type":"text","text":f"""이 이미지들은 {store_name} 카카오채널 스크린샷입니다.
모든 행사·팝업·사은혜택·이벤트를 추출하세요.
JSON 형식: {{"events":[{{"category":"상품군","name":"행사명","detail":"내용","period":"기간","type":"유형"}}]}}"""})
    resp = client.messages.create(model="claude-sonnet-4-20250514",max_tokens=2000,system=SYSTEM,messages=[{"role":"user","content":content}])
    raw = re.sub(r"```json?\n?","",resp.content[0].text).replace("```","").strip()
    return json.loads(raw).get("events",[])

def compare(lotte, hyundai):
    lt = "\n".join([f"[{e['category']}] {e['name']}: {e['detail']}" for e in lotte])
    ht = "\n".join([f"[{e['category']}] {e['name']}: {e['detail']}" for e in hyundai])
    prompt = f"""롯데백화점 대구점과 더현대 대구 행사를 비교 분석하세요.

롯데:\n{lt}\n\n더현대:\n{ht}

JSON 형식으로만 응답:
{{"categories":[{{"category":"상품군","lotte":"롯데내용","hyundai":"현대내용","winner":"롯데|더현대|비슷","point":"한줄포인트"}}],
"lotte_strength":["강점1","강점2","강점3"],
"hyundai_strength":["강점1","강점2","강점3"],
"insight":["MD 전략 제언1","제언2","제언3","제언4"]}}

상품군: 패션/스포츠·레저/뷰티/식품F&B/리빙가구/팝업스토어/사은혜택/문화이벤트
insight는 롯데 영업기획팀 입장의 실전 전략 제언으로 작성"""
    resp = client.messages.create(model="claude-sonnet-4-20250514",max_tokens=2000,messages=[{"role":"user","content":prompt}])
    raw = re.sub(r"```json?\n?","",resp.content[0].text).replace("```","").strip()
    return json.loads(raw)

@app.route("/")
def index(): return render_template_string(HTML)

@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        lotte_ev = extract_events(request.files.getlist("lotte"), "롯데백화점 대구점")
        hyundai_ev = extract_events(request.files.getlist("hyundai"), "더현대 대구")
        result = compare(lotte_ev, hyundai_ev)
        result["lotte_events"] = lotte_ev
        result["hyundai_events"] = hyundai_ev
        result["analyzed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/excel", methods=["POST"])
def excel():
    data = request.json
    wb = Workbook()
    def th(c,bg="1A1A2E",fc="FFFFFF",sz=10):
        c.fill=PatternFill("solid",start_color=bg)
        c.font=Font(bold=True,color=fc,name="맑은 고딕",size=sz)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        s=Side(border_style="thin",color="CCCCCC")
        c.border=Border(left=s,right=s,top=s,bottom=s)
    def td(c,bg="FFFFFF",fc="111111",bold=False,center=False):
        c.fill=PatternFill("solid",start_color=bg)
        c.font=Font(bold=bold,color=fc,name="맑은 고딕",size=9)
        c.alignment=Alignment(horizontal="center" if center else "left",vertical="center",wrap_text=True)
        s=Side(border_style="thin",color="E0E0E0")
        c.border=Border(left=s,right=s,top=s,bottom=s)

    # 시트1: 비교표
    ws1=wb.active; ws1.title="상품군별 비교"; ws1.sheet_view.showGridLines=False
    ws1.merge_cells("A1:E1")
    c=ws1["A1"]; c.value=f"롯데 vs 더현대 행사 비교 ({data.get('analyzed_at','')})"; th(c,sz=12)
    ws1.row_dimensions[1].height=26
    for i,(h,w) in enumerate([("상품군",14),("롯데백화점",36),("더현대 대구",36),("우세",10),("비교포인트",32)],1):
        ws1.column_dimensions[get_column_letter(i)].width=w
        th(ws1.cell(row=2,column=i,value=h))
    for r,cat in enumerate(data.get("categories",[]),3):
        ws1.row_dimensions[r].height=42
        bg="FAFAFA" if r%2==0 else "FFFFFF"
        w=cat.get("winner","비슷"); wfc="C00020" if w=="롯데" else ("003087" if w=="더현대" else "555555")
        for ci,v in enumerate([cat.get("category"),cat.get("lotte","—"),cat.get("hyundai","—"),w,cat.get("point","")],1):
            td(ws1.cell(row=r,column=ci,value=v),bg=bg,bold=(ci==1 or ci==4),fc=wfc if ci==4 else "111111",center=(ci==1 or ci==4))

    # 시트2: 롯데 상세
    ws2=wb.create_sheet("롯데 행사 상세"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:E1"); c=ws2["A1"]; c.value="롯데백화점 대구점 행사 전체"; th(c,bg="E8002D",sz=12); ws2.row_dimensions[1].height=24
    for i,(h,w) in enumerate([("카테고리",12),("행사명",28),("내용",42),("기간",14),("유형",10)],1):
        ws2.column_dimensions[get_column_letter(i)].width=w; th(ws2.cell(row=2,column=i,value=h),bg="C00020")
    for r,ev in enumerate(data.get("lotte_events",[]),3):
        bg="FFF5F5" if r%2==0 else "FFFFFF"
        for ci,k in enumerate(["category","name","detail","period","type"],1):
            td(ws2.cell(row=r,column=ci,value=ev.get(k,"")),bg=bg)
        ws2.row_dimensions[r].height=36

    # 시트3: 더현대 상세
    ws3=wb.create_sheet("더현대 행사 상세"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:E1"); c=ws3["A1"]; c.value="더현대 대구 행사 전체"; th(c,bg="003087",sz=12); ws3.row_dimensions[1].height=24
    for i,(h,w) in enumerate([("카테고리",12),("행사명",28),("내용",42),("기간",14),("유형",10)],1):
        ws3.column_dimensions[get_column_letter(i)].width=w; th(ws3.cell(row=2,column=i,value=h),bg="003087")
    for r,ev in enumerate(data.get("hyundai_events",[]),3):
        bg="F0F4FF" if r%2==0 else "FFFFFF"
        for ci,k in enumerate(["category","name","detail","period","type"],1):
            td(ws3.cell(row=r,column=ci,value=ev.get(k,"")),bg=bg)
        ws3.row_dimensions[r].height=36

    # 시트4: AI 분석
    ws4=wb.create_sheet("AI 분석"); ws4.sheet_view.showGridLines=False
    ws4.column_dimensions["A"].width=18; ws4.column_dimensions["B"].width=60
    ws4.merge_cells("A1:B1"); c=ws4["A1"]; c.value="AI 경쟁 분석 — MD 전략 제언"; th(c,sz=12); ws4.row_dimensions[1].height=24
    row=2
    for title,items,bg,fc in [("롯데 강점",data.get("lotte_strength",[]),"FCE4E4","C00020"),
                                ("더현대 강점",data.get("hyundai_strength",[]),"E4EBF9","003087"),
                                ("MD 전략 제언",data.get("insight",[]),"FFFDE7","856404")]:
        th(ws4.cell(row=row,column=1,value=title),bg=fc)
        ws4.cell(row=row,column=2).fill=PatternFill("solid",start_color=fc)
        ws4.row_dimensions[row].height=20; row+=1
        for item in items:
            td(ws4.cell(row=row,column=2,value=f"• {item}"),bg=bg)
            ws4.cell(row=row,column=1).fill=PatternFill("solid",start_color=bg)
            ws4.row_dimensions[row].height=30; row+=1
        row+=1

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,
                     download_name=f"백화점행사비교_{datetime.now().strftime('%Y%m%d')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__":
    print("="*50)
    print("  백화점 행사 AI 분석 시스템")
    print("  http://localhost:5000")
    print("="*50)
    app.run(debug=True,port=5000)
